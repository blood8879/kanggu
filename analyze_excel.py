import pandas as pd
import openpyxl
import re
from openpyxl.utils import get_column_letter
import json

def analyze_excel_file(file_path):
    """
    엑셀 파일을 분석하여 input1, input2, input3... 형태의 패턴을 찾아 위치를 반환
    """
    result = {
        "sheets": {},
        "input_fields": [],
        "file_info": {}
    }
    
    try:
        # openpyxl로 워크북 열기
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        
        result["file_info"]["sheet_names"] = workbook.sheetnames
        result["file_info"]["total_sheets"] = len(workbook.sheetnames)
        
        # 각 시트 분석
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = {
                "name": sheet_name,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "input_fields": [],
                "all_values": []
            }
            
            # 모든 셀 검사
            for row in range(1, sheet.max_row + 1):
                row_data = []
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_value = cell.value
                    
                    # 셀 값을 문자열로 변환
                    if cell_value is not None:
                        cell_str = str(cell_value)
                        row_data.append(cell_str)
                        
                        # input 패턴 검사 (input1, input2, input3...)
                        input_pattern = re.search(r'input\d+', cell_str, re.IGNORECASE)
                        if input_pattern:
                            input_field = {
                                "pattern": input_pattern.group(),
                                "cell": f"{get_column_letter(col)}{row}",
                                "row": row,
                                "column": col,
                                "column_letter": get_column_letter(col),
                                "original_value": cell_str,
                                "sheet": sheet_name
                            }
                            sheet_data["input_fields"].append(input_field)
                            result["input_fields"].append(input_field)
                    else:
                        row_data.append("")
                
                sheet_data["all_values"].append(row_data)
            
            result["sheets"][sheet_name] = sheet_data
        
        # pandas로도 데이터 확인
        try:
            df = pd.read_excel(file_path, sheet_name=None)
            result["pandas_info"] = {}
            for sheet_name, data in df.items():
                result["pandas_info"][sheet_name] = {
                    "shape": data.shape,
                    "columns": data.columns.tolist(),
                    "dtypes": data.dtypes.to_dict()
                }
        except Exception as e:
            result["pandas_error"] = str(e)
        
        return result
        
    except Exception as e:
        return {"error": str(e)}

def print_analysis_result(result):
    """
    분석 결과를 보기 좋게 출력
    """
    if "error" in result:
        print(f"❌ 오류 발생: {result['error']}")
        return
    
    print("🔍 엑셀 파일 분석 결과")
    print("=" * 50)
    
    # 파일 정보
    print(f"📄 총 시트 수: {result['file_info']['total_sheets']}")
    print(f"📋 시트 이름: {', '.join(result['file_info']['sheet_names'])}")
    print()
    
    # Input 필드 정보
    print(f"🎯 발견된 Input 필드 수: {len(result['input_fields'])}")
    if result['input_fields']:
        print("Input 필드 상세:")
        for i, field in enumerate(result['input_fields'], 1):
            print(f"  {i}. {field['pattern']} - {field['sheet']}!{field['cell']} - '{field['original_value']}'")
    else:
        print("  ⚠️  input 패턴을 찾을 수 없습니다.")
    print()
    
    # 각 시트 정보
    for sheet_name, sheet_data in result['sheets'].items():
        print(f"📊 시트 '{sheet_name}':")
        print(f"  크기: {sheet_data['max_row']}행 × {sheet_data['max_column']}열")
        print(f"  Input 필드: {len(sheet_data['input_fields'])}개")
        
        # 시트의 첫 몇 행 데이터 샘플 출력
        if sheet_data['all_values']:
            print("  데이터 샘플 (첫 5행):")
            for i, row in enumerate(sheet_data['all_values'][:5], 1):
                # 빈 값들을 제거하고 출력
                non_empty = [cell for cell in row if cell.strip()]
                if non_empty:
                    print(f"    행 {i}: {' | '.join(non_empty[:5])}")
        print()

if __name__ == "__main__":
    file_path = "sample.xlsx"
    print("🚀 엑셀 파일 분석 시작...")
    print()
    
    result = analyze_excel_file(file_path)
    print_analysis_result(result)
    
    # JSON 파일로 저장
    with open("excel_analysis.json", "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print("💾 분석 결과가 'excel_analysis.json'에 저장되었습니다.")