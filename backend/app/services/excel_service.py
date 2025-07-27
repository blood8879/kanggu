import pandas as pd
import openpyxl
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.cell.rich_text import TextBlock, CellRichText
import os
import uuid
import shutil
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, List, Optional
from app.models.excel_models import ExcelAnalysisResult, InputField, SheetInfo, InputValue

class ExcelService:
    def __init__(self, upload_dir: str = "uploads"):
        self.upload_dir = upload_dir
        os.makedirs(upload_dir, exist_ok=True)
    
    def _replace_pattern_preserving_format(self, cell, pattern: str, replacement: str):
        """서식을 보존하면서 패턴을 교체하는 헬퍼 메소드"""
        try:
            original_value = cell.value
            
            # CellRichText 타입인지 확인
            if isinstance(original_value, CellRichText):
                # Rich Text의 각 부분을 순회하면서 패턴 교체
                new_rich_text = CellRichText()
                
                for part in original_value:
                    if hasattr(part, 'text') and hasattr(part, 'font'):
                        # TextBlock인 경우
                        new_text = re.sub(
                            re.escape(pattern), 
                            replacement, 
                            part.text, 
                            flags=re.IGNORECASE
                        )
                        # 원래 폰트 서식을 유지하면서 새 텍스트 블록 생성
                        from openpyxl.cell.rich_text import TextBlock
                        new_rich_text.append(TextBlock(new_text, part.font))
                    else:
                        # 문자열인 경우
                        text_str = str(part)
                        new_text = re.sub(
                            re.escape(pattern), 
                            replacement, 
                            text_str, 
                            flags=re.IGNORECASE
                        )
                        new_rich_text.append(new_text)
                
                cell.value = new_rich_text
                
            else:
                # 일반 텍스트 처리
                original_value_str = str(original_value)
                new_value = re.sub(
                    re.escape(pattern), 
                    replacement, 
                    original_value_str, 
                    flags=re.IGNORECASE
                )
                cell.value = new_value
                
        except Exception as e:
            # 실패시 기본 방식으로 폴백
            print(f"Pattern replacement failed for cell, using fallback: {e}")
            original_value_str = str(cell.value) if cell.value else ""
            new_value = re.sub(
                re.escape(pattern), 
                replacement, 
                original_value_str, 
                flags=re.IGNORECASE
            )
            cell.value = new_value
    
    def _process_excel_via_xml(self, original_file_path: str, output_path: str, input_values: List[InputValue], processed_file_id: str) -> str:
        """XML 기반으로 Excel 파일 처리하여 서식 완전 보존"""
        # 원본 파일을 복사
        shutil.copy2(original_file_path, output_path)
        
        # Excel 파일은 실제로는 ZIP 아카이브
        temp_dir = os.path.join(self.upload_dir, f"temp_{processed_file_id}")
        os.makedirs(temp_dir, exist_ok=True)
        
        try:
            # ZIP으로 압축 해제
            with zipfile.ZipFile(output_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # worksheet XML 파일에서 텍스트 교체
            worksheet_path = os.path.join(temp_dir, 'xl', 'worksheets', 'sheet1.xml')
            if os.path.exists(worksheet_path):
                with open(worksheet_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                # 패턴별로 교체
                for input_value in input_values:
                    pattern = input_value.pattern
                    replacement = input_value.value
                    # XML 내에서 텍스트 교체 (대소문자 구분 없이)
                    content = re.sub(
                        re.escape(pattern), 
                        replacement, 
                        content, 
                        flags=re.IGNORECASE
                    )
                
                with open(worksheet_path, 'w', encoding='utf-8') as f:
                    f.write(content)
            
            # sharedStrings.xml도 처리 (공유 문자열 테이블)
            shared_strings_path = os.path.join(temp_dir, 'xl', 'sharedStrings.xml')
            if os.path.exists(shared_strings_path):
                with open(shared_strings_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                for input_value in input_values:
                    pattern = input_value.pattern
                    replacement = input_value.value
                    content = re.sub(
                        re.escape(pattern), 
                        replacement, 
                        content, 
                        flags=re.IGNORECASE
                    )
                
                with open(shared_strings_path, 'w', encoding='utf-8') as f:
                    f.write(content)
            
            # 다시 ZIP으로 압축
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arc_name = os.path.relpath(file_path, temp_dir)
                        zip_ref.write(file_path, arc_name)
            
            return processed_file_id
            
        finally:
            # 임시 디렉토리 정리
            shutil.rmtree(temp_dir, ignore_errors=True)
    
    def _process_excel_via_openpyxl(self, original_file_path: str, input_values: List[InputValue]) -> str:
        """openpyxl을 사용한 기본 처리 방식"""
        processed_file_id = str(uuid.uuid4())
        output_path = os.path.join(self.upload_dir, f"processed_{processed_file_id}.xlsx")
        
        # 원본 파일을 먼저 복사
        shutil.copy2(original_file_path, output_path)
        
        # 복사된 파일을 열어서 값만 수정
        workbook = openpyxl.load_workbook(output_path, data_only=False)
        
        for input_value in input_values:
            sheet = workbook[input_value.sheet]
            
            cell_match = re.match(r'([A-Z]+)(\d+)', input_value.cell)
            if cell_match:
                col_letter, row_num = cell_match.groups()
                row = int(row_num)
                col = openpyxl.utils.column_index_from_string(col_letter)
                
                cell = sheet.cell(row=row, column=col)
                
                if cell.value is not None:
                    self._replace_pattern_preserving_format(cell, input_value.pattern, input_value.value)
                else:
                    cell.value = input_value.value
        
        workbook.save(output_path)
        workbook.close()
        
        return processed_file_id
    
    def analyze_excel_file(self, file_path: str) -> ExcelAnalysisResult:
        result = {
            "sheets": {},
            "input_fields": [],
            "file_info": {}
        }
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            
            result["file_info"]["sheet_names"] = workbook.sheetnames
            result["file_info"]["total_sheets"] = len(workbook.sheetnames)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_input_fields = []
                
                for row in range(1, sheet.max_row + 1):
                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row, column=col)
                        cell_value = cell.value
                        
                        if cell_value is not None:
                            cell_str = str(cell_value)
                            
                            input_pattern = re.search(r'input\d+', cell_str, re.IGNORECASE)
                            if input_pattern:
                                input_field = InputField(
                                    pattern=input_pattern.group(),
                                    cell=f"{get_column_letter(col)}{row}",
                                    row=row,
                                    column=col,
                                    column_letter=get_column_letter(col),
                                    original_value=cell_str,
                                    sheet=sheet_name
                                )
                                sheet_input_fields.append(input_field)
                                result["input_fields"].append(input_field)
                
                sheet_info = SheetInfo(
                    name=sheet_name,
                    max_row=sheet.max_row,
                    max_column=sheet.max_column,
                    input_fields=sheet_input_fields
                )
                result["sheets"][sheet_name] = sheet_info
            
            try:
                df = pd.read_excel(file_path, sheet_name=None)
                pandas_info = {}
                for sheet_name, data in df.items():
                    pandas_info[sheet_name] = {
                        "shape": data.shape,
                        "columns": data.columns.tolist(),
                        "dtypes": {k: str(v) for k, v in data.dtypes.to_dict().items()}
                    }
                result["pandas_info"] = pandas_info
            except Exception as e:
                result["pandas_error"] = str(e)
            
            return ExcelAnalysisResult(**result)
            
        except Exception as e:
            raise Exception(f"Excel 분석 중 오류 발생: {str(e)}")
    
    def process_excel_with_inputs(self, original_file_path: str, input_values: List[InputValue]) -> str:
        try:
            processed_file_id = str(uuid.uuid4())
            output_path = os.path.join(self.upload_dir, f"processed_{processed_file_id}.xlsx")
            
            # XML 기반 처리로 서식 완전 보존
            return self._process_excel_via_xml(original_file_path, output_path, input_values, processed_file_id)
            
        except Exception as e:
            # XML 처리 실패시 기본 방식으로 폴백
            print(f"XML processing failed, falling back to openpyxl: {e}")
            return self._process_excel_via_openpyxl(original_file_path, input_values)
    
    def process_sample_excel_with_inputs(self, input_values: List[InputValue]) -> str:
        """sample.xlsx 파일을 기반으로 입력값을 처리합니다."""
        try:
            # 프로젝트 루트에 있는 sample.xlsx 파일 경로
            # backend/app/services/excel_service.py -> ../../../sample.xlsx
            current_dir = os.path.dirname(os.path.abspath(__file__))
            backend_dir = os.path.dirname(os.path.dirname(current_dir))
            project_root = os.path.dirname(backend_dir)
            sample_file_path = os.path.join(project_root, "sample.xlsx")
            
            if not os.path.exists(sample_file_path):
                raise Exception("sample.xlsx 파일을 찾을 수 없습니다.")
            
            return self.process_excel_with_inputs(sample_file_path, input_values)
            
        except Exception as e:
            raise Exception(f"Sample Excel 처리 중 오류 발생: {str(e)}")
    
    def get_file_path(self, file_id: str, prefix: str = "") -> str:
        filename = f"{prefix}{file_id}.xlsx"
        return os.path.join(self.upload_dir, filename)
    
    def save_uploaded_file(self, file_content: bytes, filename: str) -> str:
        file_id = str(uuid.uuid4())
        file_path = self.get_file_path(file_id, "upload_")
        
        with open(file_path, "wb") as f:
            f.write(file_content)
        
        return file_id